import os
from io import BytesIO
from pathlib import Path

import base64
import pandas as pd
import streamlit as st
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

try:
    import cloudinary
    import cloudinary.uploader
    CLOUDINARY_AVAILABLE = True
except ImportError:
    CLOUDINARY_AVAILABLE = False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

st.set_page_config(page_title="Quản lý hàng hóa", page_icon="📦", layout="wide")

CREDENTIALS = {
    "purchasing": {"password": "1", "role": "PURCHASING"},
    "sale": {"password": "1", "role": "SALE"},
}

IMG_DIR = Path("images")
IMG_DIR.mkdir(parents=True, exist_ok=True)
DATA_DIR = Path("data")
DATA_DIR.mkdir(parents=True, exist_ok=True)
DEFAULT_COLUMNS = ["Tên hàng", "Code", "Phân loại", "CBM", "Packing", "Giá", "Ảnh"]

USE_GOOGLE_SHEETS = False
worksheet = None
sheet_status_message = ""

sheet_id = None
if "sheet_id" in st.secrets:
    sheet_id = st.secrets["sheet_id"]
elif "gcp_service_account" in st.secrets:
    sheet_id = st.secrets["gcp_service_account"].get("sheet_id")

if GSPREAD_AVAILABLE and "gcp_service_account" in st.secrets and sheet_id:
    try:
        service_account_info = dict(st.secrets["gcp_service_account"])
        if "private_key" in service_account_info:
            pk = service_account_info["private_key"]
            if "\\n" in pk:
                pk = pk.replace("\\n", "\n")
            service_account_info["private_key"] = pk

        credentials = Credentials.from_service_account_info(
            service_account_info,
            scopes=["https://www.googleapis.com/auth/spreadsheets"],
        )
        gc = gspread.authorize(credentials)
        SHEET_ID = sheet_id
        sh = gc.open_by_key(SHEET_ID)
        try:
            worksheet = sh.worksheet("Products")
        except Exception:
            worksheet = sh.add_worksheet("Products", rows=1000, cols=len(DEFAULT_COLUMNS))
            worksheet.append_row(DEFAULT_COLUMNS)
        USE_GOOGLE_SHEETS = True
        sheet_status_message = "✅ Kết nối Google Sheets thành công. Dữ liệu sẽ được lưu trên cloud."
    except Exception as e:
        USE_GOOGLE_SHEETS = False
        sheet_status_message = f"⚠️ Không thể kết nối Google Sheets: {e}"
else:
    if not GSPREAD_AVAILABLE:
        sheet_status_message = "⚠️ Thiếu thư viện gspread/google-auth. Cài đặt bằng pip nếu muốn dùng Google Sheets."
    elif "gcp_service_account" not in st.secrets:
        sheet_status_message = "⚠️ Thiếu st.secrets['gcp_service_account'] trong .streamlit/secrets.toml."
    elif not sheet_id:
        sheet_status_message = "⚠️ Thiếu st.secrets['sheet_id'] trong .streamlit/secrets.toml. Nếu bạn đặt sheet_id trong bảng gcp_service_account, app cũng sẽ đọc được." 
    else:
        sheet_status_message = "⚠️ Google Sheets không khả dụng. Sử dụng CSV local."

if CLOUDINARY_AVAILABLE and "cloudinary" in st.secrets:
    cloudinary.config(
        cloud_name=st.secrets["cloudinary"].get("cloud_name"),
        api_key=st.secrets["cloudinary"].get("api_key"),
        api_secret=st.secrets["cloudinary"].get("api_secret")
    )


def load_data():
    if USE_GOOGLE_SHEETS and worksheet is not None:
        try:
            data = worksheet.get_all_records()
            if not data:
                return pd.DataFrame(columns=DEFAULT_COLUMNS)
            df = pd.DataFrame(data)
            for col in DEFAULT_COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            return df[DEFAULT_COLUMNS]
        except Exception as e:
            st.error(f"Lỗi khi tải từ Google Sheets: {e}")
            return pd.DataFrame(columns=DEFAULT_COLUMNS)
    else:
        csv_path = DATA_DIR / "database.csv"
        if csv_path.exists():
            df = pd.read_csv(csv_path)
        else:
            df = pd.DataFrame(columns=DEFAULT_COLUMNS)
        for col in DEFAULT_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[DEFAULT_COLUMNS]


def save_data(df):
    if USE_GOOGLE_SHEETS and worksheet is not None:
        try:
            worksheet.clear()
            worksheet.append_row(DEFAULT_COLUMNS)
            for _, row in df.iterrows():
                worksheet.append_row(row.tolist())
            return True, "✅ Dữ liệu đã được lưu vào Google Sheets."
        except Exception as e:
            return False, f"❌ Lỗi khi lưu vào Google Sheets: {e}"
    else:
        try:
            csv_path = DATA_DIR / "database.csv"
            df.to_csv(csv_path, index=False)
            return True, "✅ Dữ liệu đã được lưu vào CSV local."
        except Exception as e:
            return False, f"❌ Lỗi khi lưu CSV local: {e}"


def upload_to_cloudinary(image_file, code):
    """Upload image to Cloudinary and return secure_url."""
    if not CLOUDINARY_AVAILABLE or "cloudinary" not in st.secrets:
        return None
    try:
        result = cloudinary.uploader.upload(
            image_file,
            public_id=f"khai_nam_{code}",
            overwrite=True,
            resource_type="auto"
        )
        return result.get("secure_url")
    except Exception as e:
        st.error(f"❌ Lỗi upload ảnh lên Cloudinary: {e}")
        return None


def categories_list():
    df = load_data()
    categories = [c for c in df["Phân loại"].dropna().astype(str).unique() if c.strip()]
    defaults = ["Nước uống", "Bún khô", "Gia vị", "Đồ hộp", "Khác"]
    for item in defaults:
        if item not in categories:
            categories.insert(0, item)
    return categories


def save_product(name, code, category, cbm, packing, price, image_file):
    df = load_data()
    code = str(code).strip()
    if code in df["Code"].astype(str).values:
        return False, "Mã Code đã tồn tại. Vui lòng dùng mã khác hoặc cập nhật bản ghi hiện tại."

    try:
        image = Image.open(image_file)
        image = image.convert("RGB")
    except Exception:
        return False, "Không thể đọc ảnh. Vui lòng thử lại với ảnh hợp lệ."

    image_url = upload_to_cloudinary(image_file, code)
    if not image_url:
        image_url = ""

    new_row = {
        "Tên hàng": name.strip(),
        "Code": code,
        "Phân loại": category.strip(),
        "CBM": cbm.strip(),
        "Packing": packing.strip(),
        "Giá": price.strip(),
        "Ảnh": image_url,
    }
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    return save_data(df)


def update_product(code, name, category, cbm, packing, price, image_file=None):
    df = load_data()
    index = df.index[df["Code"].astype(str) == code].tolist()
    if not index:
        return False, "Không tìm thấy hàng hóa để cập nhật."

    idx = index[0]
    df.at[idx, "Tên hàng"] = name.strip()
    df.at[idx, "Phân loại"] = category.strip()
    df.at[idx, "CBM"] = cbm.strip()
    df.at[idx, "Packing"] = packing.strip()
    df.at[idx, "Giá"] = price.strip()

    if image_file is not None:
        try:
            image = Image.open(image_file)
            image = image.convert("RGB")
            image_url = upload_to_cloudinary(image_file, code)
            if image_url:
                df.at[idx, "Ảnh"] = image_url
        except Exception:
            return False, "Không thể cập nhật ảnh. Vui lòng thử lại với ảnh hợp lệ."

    return save_data(df)


def delete_product(code):
    df = load_data()
    df = df[df["Code"].astype(str) != code].reset_index(drop=True)
    return save_data(df)


def authenticate(username, password):
    user = CREDENTIALS.get(username)
    if user and user["password"] == password:
        return user["role"]
    return None


def create_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "PO"

    headers = list(df.columns)
    ws.append(headers)

    for row in df.itertuples(index=False):
        values = []
        for header, value in zip(headers, row):
            if header == "Ảnh":
                if pd.notna(value) and str(value).startswith("http"):
                    values.append(str(value))
                elif pd.notna(value) and value:
                    values.append(Path(value).name)
                else:
                    values.append("")
            else:
                values.append(value)
        ws.append(values)

    for col_idx, header in enumerate(headers, start=1):
        width = 30 if header == "Ảnh" else 18
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    image_col = headers.index("Ảnh") + 1
    for row_idx, image_path in enumerate(df["Ảnh"], start=2):
        if pd.notna(image_path) and str(image_path).strip():
            if not str(image_path).startswith("http"):
                image_file = Path(image_path)
                if image_file.exists():
                    try:
                        img = XLImage(str(image_file))
                        img.width = 120
                        img.height = 120
                        img.anchor = f"{get_column_letter(image_col)}{row_idx}"
                        ws.add_image(img)
                        ws.row_dimensions[row_idx].height = 90
                    except Exception:
                        continue

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer



def get_image_html(img_val):
    if not pd.notna(img_val) or not str(img_val).strip():
        return "https://via.placeholder.com/300x200?text=No+Image"
    img_val = str(img_val).strip()
    if img_val.startswith("http"):
        return img_val
    else:
        path = Path(img_val)
        if path.exists():
            try:
                with open(path, "rb") as f:
                    encoded = base64.b64encode(f.read()).decode()
                mime = "image/jpeg"
                if path.suffix.lower() == ".png":
                    mime = "image/png"
                elif path.suffix.lower() == ".webp":
                    mime = "image/webp"
                return f"data:{mime};base64,{encoded}"
            except:
                pass
        return "https://via.placeholder.com/300x200?text=Not+Found"

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
    st.session_state["username"] = ""
    st.session_state["role"] = ""
    st.session_state["edit_code"] = ""
    st.session_state["current_page_purchasing"] = 0
    st.session_state["current_page_sale"] = 0

if "login_error" not in st.session_state:
    st.session_state["login_error"] = ""

st.markdown(
    """
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap');
        
        * {
            font-family: 'Inter', 'Segoe UI', sans-serif;
            margin: 0;
            padding: 0;
        }

        .logo-container {
            text-align: center;
            padding: 2.5rem 0;
            background: linear-gradient(135deg, #4F46E5 0%, #312E81 100%);
            border-radius: 16px;
            color: white;
            margin-bottom: 2rem;
            box-shadow: 0 10px 25px -5px rgba(79, 70, 229, 0.4);
        }

        .logo-text {
            font-size: 52px;
            font-weight: 800;
            letter-spacing: 2px;
            margin: 1rem 0 0.5rem 0;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }

        .logo-tagline {
            font-size: 18px;
            font-style: italic;
            color: #FCD34D;
            font-weight: 600;
        }

        .header-info {
            background-color: #F8FAFC;
            padding: 1.2rem 1.5rem;
            border-radius: 12px;
            border-left: 4px solid #4F46E5;
            margin-bottom: 1.5rem;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
        }

        .stButton>button {
            background-color: #4F46E5;
            color: white;
            border-radius: 8px;
            font-weight: 600;
            border: none;
            padding: 0.6rem 1.2rem;
            transition: all 0.3s ease;
        }

        .stButton>button:hover {
            background-color: #4338CA;
            color: white;
            box-shadow: 0 4px 12px rgba(79, 70, 229, 0.3);
            transform: translateY(-2px);
        }

        .stTextInput>div>div>input,
        .stSelectbox>div>div>div>select,
        .stTextArea>div>div>textarea {
            border-radius: 8px;
            border: 1px solid #CBD5E1;
            padding: 0.8rem;
            transition: all 0.2s ease;
        }

        .stTextInput>div>div>input:focus,
        .stSelectbox>div>div>div>select:focus,
        .stTextArea>div>div>textarea:focus {
            border: 2px solid #4F46E5;
            box-shadow: 0 0 0 3px rgba(79, 70, 229, 0.2);
        }

        .stFileUploader>div>div {
            border: 2px dashed #818CF8;
            border-radius: 12px;
            padding: 2rem;
            background-color: #EEF2FF;
            transition: all 0.3s ease;
        }

        h1, h2, h3, h4 {
            color: #1E293B;
            font-weight: 800;
        }

        .info-box {
            background-color: #FEF3C7;
            padding: 1.2rem;
            border-radius: 12px;
            border-left: 4px solid #F59E0B;
            margin-bottom: 1.5rem;
            color: #92400E;
            font-weight: 500;
        }
        
        /* Grid Card Hover Effects */
        [data-testid="stVerticalBlockBorderWrapper"] {
            border-radius: 16px !important;
            transition: all 0.3s ease !important;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05) !important;
            background-color: white !important;
            border: 1px solid #E2E8F0 !important;
        }
        [data-testid="stVerticalBlockBorderWrapper"]:hover {
            transform: translateY(-5px) !important;
            box-shadow: 0 15px 25px -5px rgba(0, 0, 0, 0.1), 0 8px 10px -6px rgba(0, 0, 0, 0.1) !important;
            border-color: #818CF8 !important;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

if not st.session_state["logged_in"]:
    st.markdown(
        """
        <div class="logo-container">
            <div style="font-size: 60px; margin-bottom: 0.5rem;">⛵🌍✈️</div>
            <div class="logo-text">KHAI NAM TRADING</div>
            <div class="logo-tagline">"Connecting the World – Growing Together"</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("""---""")
    st.markdown(f"<div class='info-box'>{sheet_status_message}</div>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("### 🔐 Đăng nhập hệ thống")
        with st.form("login_form"):
            username = st.text_input("👤 Username")
            password = st.text_input("🔑 Password", type="password")
            submitted = st.form_submit_button("🚀 Đăng nhập", use_container_width=True)
            if submitted:
                role = authenticate(username, password)
                if role:
                    st.session_state["logged_in"] = True
                    st.session_state["username"] = username
                    st.session_state["role"] = role
                    st.session_state["login_error"] = ""
                    st.rerun()
                else:
                    st.session_state["login_error"] = "Tên đăng nhập hoặc mật khẩu không đúng."
        if st.session_state["login_error"]:
            st.error(st.session_state["login_error"])
    st.stop()

is_purchasing = st.session_state["role"] == "PURCHASING"

st.markdown(
    """
    <div class="logo-container">
        <div style="font-size: 40px; margin-bottom: 0.3rem;">⛵🌍✈️</div>
        <div class="logo-text" style="font-size: 36px; margin: 0.5rem 0 0.3rem 0;">KHAI NAM TRADING</div>
        <div class="logo-tagline">"Connecting the World – Growing Together"</div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown("""---""")
header_col, user_col, logout_col = st.columns([2, 2, 1])
with header_col:
    st.markdown("### 📦 Quản lý hàng hóa")
with user_col:
    role_badge = "🎯 THU MUA" if is_purchasing else "💼 SALE"
    st.markdown(f"<div class='header-info'><b>Vai trò:</b> {role_badge}<br><b>Người dùng:</b> {st.session_state['username']}<br><b>Dữ liệu:</b> {'Google Sheets' if USE_GOOGLE_SHEETS else 'CSV Local'}</div>", unsafe_allow_html=True)
with logout_col:
    if st.button("🚪 Đăng xuất", use_container_width=True):
        st.session_state["logged_in"] = False
        st.session_state["username"] = ""
        st.session_state["role"] = ""
        st.session_state["edit_code"] = ""
        st.rerun()

st.markdown("""---""")

tabs = st.tabs(["📝 Nhập hàng", "📋 Danh sách hàng"] if is_purchasing else ["📋 Danh sách hàng"])

if is_purchasing:
    with tabs[0]:
        st.markdown("### 📝 Nhập thông tin hàng hóa")
        st.markdown('<div class="info-box">⚡ Vui lòng nhập đầy đủ thông tin sản phẩm để thêm vào danh sách.</div>', unsafe_allow_html=True)
        categories = categories_list()
        with st.form("form_add_product", clear_on_submit=True):
            with st.container():
                col1, col2 = st.columns(2)
                name = col1.text_input("📦 Tên hàng")
                code = col2.text_input("🏷️ Code")
                category = col1.selectbox("📂 Phân loại", options=categories + ["Nhập mới..."])
                if category == "Nhập mới...":
                    category = col2.text_input("📝 Phân loại mới")
                cbm = col1.text_input("📏 CBM")
                packing = col2.text_input("📦 Packing")
                price = col1.text_input("💰 Giá")
                image_file = col2.file_uploader("📸 Upload ảnh", type=["jpg", "jpeg", "png", "webp"])

            if image_file is not None:
                st.image(image_file, caption="✅ Ảnh đã chọn", width=300)

            submit = st.form_submit_button("💾 Lưu hàng", use_container_width=True)
            if submit:
                if not all([name.strip(), code.strip(), category.strip(), cbm.strip(), packing.strip(), price.strip(), image_file]):
                    st.error("❌ Vui lòng nhập đầy đủ thông tin và chọn ảnh.")
                else:
                    success, message = save_product(name, code, category, cbm, packing, price, image_file)
                    if success:
                        st.success(message)
                        st.rerun()
                    else:
                        st.error(message)
    list_tab = tabs[1]
else:
    list_tab = tabs[0]

with list_tab:
    st.markdown("### 📋 Danh sách hàng hóa")
    df = load_data()
    if df.empty:
        st.info("ℹ️ Chưa có mặt hàng nào. Vui lòng thêm hàng hóa nếu bạn là Thu mua.")
    else:
        selected_rows = []
        if is_purchasing:
            edit_code = st.session_state.get("edit_code", "")
            if edit_code:
                edit_df = df[df["Code"].astype(str) == edit_code]
                if not edit_df.empty:
                    row = edit_df.iloc[0]
                    st.markdown("---")
                    st.markdown(f"### ✏️ Chỉnh sửa hàng hóa: {row['Tên hàng']} ({row['Code']})")
                    with st.form("edit_form", clear_on_submit=False):
                        col1, col2 = st.columns(2)
                        edit_name = col1.text_input("📦 Tên hàng", value=row["Tên hàng"])
                        edit_category = col2.selectbox(
                            "📂 Phân loại",
                            options=categories_list() + ["Nhập mới..."],
                            index=0 if row["Phân loại"] in categories_list() else len(categories_list()),
                        )
                        if edit_category == "Nhập mới...":
                            edit_category = col1.text_input("📝 Phân loại mới", value=row["Phân loại"])
                        edit_cbm = col1.text_input("📏 CBM", value=row["CBM"])
                        edit_packing = col2.text_input("📦 Packing", value=row["Packing"])
                        edit_price = col1.text_input("💰 Giá", value=row["Giá"])
                        edit_image = col2.file_uploader("📸 Cập nhật ảnh (không bắt buộc)", type=["jpg", "jpeg", "png", "webp"])

                        save_edit = st.form_submit_button("✅ Cập nhật")
                        cancel_edit = st.form_submit_button("❌ Hủy")
                        if save_edit:
                            success, message = update_product(edit_code, edit_name, edit_category, edit_cbm, edit_packing, edit_price, edit_image if edit_image else None)
                            if success:
                                st.success(message)
                                st.session_state["edit_code"] = ""
                                st.rerun()
                            else:
                                st.error(message)
                        if cancel_edit:
                            st.session_state["edit_code"] = ""
                            st.rerun()
                else:
                    st.session_state["edit_code"] = ""

        # --- SMART FILTERING TOP BAR ---
        st.markdown("---")
        with st.container():
            filter_cols = st.columns([1.5, 2])
            with filter_cols[0]:
                search_q = st.text_input("🔍 Tìm kiếm (Tên / Code)", key="search_q")
            with filter_cols[1]:
                cats = categories_list()
                selected_cats = st.multiselect("📂 Phân loại", options=cats, key="cat_q")
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        with st.spinner("Đang tải danh sách..."):
            filtered_df = df.copy()
            if search_q:
                q = search_q.lower()
                filtered_df = filtered_df[
                    filtered_df["Tên hàng"].astype(str).str.lower().str.contains(q) |
                    filtered_df["Code"].astype(str).str.lower().str.contains(q)
                ]
            if selected_cats:
                filtered_df = filtered_df[filtered_df["Phân loại"].isin(selected_cats)]

            if filtered_df.empty:
                st.info("ℹ️ Không có mặt hàng nào khớp với tìm kiếm của bạn.")
            else:
                items_per_page = 20
                total_pages = (len(filtered_df) + items_per_page - 1) // items_per_page
                
                # Check pagination bounds
                current_page_key = "current_page_purchasing" if is_purchasing else "current_page_sale"
                if st.session_state[current_page_key] >= total_pages:
                    st.session_state[current_page_key] = max(0, total_pages - 1)
                
                start_idx = st.session_state[current_page_key] * items_per_page
                end_idx = start_idx + items_per_page
                df_page = filtered_df.iloc[start_idx:end_idx]
                
                # --- GRID LAYOUT ---
                cols_per_row = 4
                for i in range(0, len(df_page), cols_per_row):
                    grid_cols = st.columns(cols_per_row)
                    for j in range(cols_per_row):
                        if i + j < len(df_page):
                            row = df_page.iloc[i + j]
                            with grid_cols[j]:
                                with st.container(border=True):
                                    img_src = get_image_html(row["Ảnh"])
                                    st.markdown(f'''
                                    <div style="text-align: center;">
                                        <img src="{img_src}" style="width: 100%; height: 200px; object-fit: cover; border-radius: 8px; margin-bottom: 10px;">
                                    </div>
                                    <div style="padding: 5px;">
                                        <h4 style="margin: 0; color: #1E293B; font-size: 1.1rem; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;" title="{row['Tên hàng']}">{row['Tên hàng']}</h4>
                                        <p style="margin: 5px 0 0 0; color: #64748B; font-size: 0.9rem;">🏷️ <b>Code:</b> {row['Code']}</p>
                                        <p style="margin: 2px 0 0 0; color: #64748B; font-size: 0.9rem;">📂 {row['Phân loại']}</p>
                                        <p style="margin: 5px 0 10px 0; color: #4F46E5; font-weight: 700; font-size: 1.1rem;">💰 {row['Giá']}</p>
                                    </div>
                                    ''', unsafe_allow_html=True)
                                    
                                    selected = st.checkbox("✅ Chọn", key=f"select_{row['Code']}")
                                    if selected:
                                        selected_rows.append(row)
                                        
                                    if is_purchasing:
                                        action_cols = st.columns(2)
                                        if action_cols[0].button("✏️", key=f"edit_{row['Code']}", use_container_width=True):
                                            st.session_state["edit_code"] = str(row["Code"])
                                            st.rerun()
                                        if action_cols[1].button("🗑️", key=f"delete_{row['Code']}", use_container_width=True):
                                            success, message = delete_product(str(row["Code"]))
                                            if success:
                                                st.success(message)
                                            else:
                                                st.error(message)
                                            st.rerun()

                # Pagination UI
                st.markdown("<br>", unsafe_allow_html=True)
                col1, col2, col3 = st.columns([1, 2, 1])
                with col1:
                    if st.button("⬅️ Trang trước", use_container_width=True, disabled=st.session_state[current_page_key] == 0):
                        st.session_state[current_page_key] -= 1
                        st.rerun()
                with col2:
                    st.markdown(f"<div style='text-align: center; padding: 0.5rem; background: #EEF2FF; border-radius: 8px; color: #4F46E5; font-weight: 600;'>Trang {st.session_state[current_page_key] + 1} / {total_pages}</div>", unsafe_allow_html=True)
                with col3:
                    if st.button("Trang sau ➡️", use_container_width=True, disabled=st.session_state[current_page_key] >= total_pages - 1):
                        st.session_state[current_page_key] += 1
                        st.rerun()

        # Excel Export Bottom
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("📊 Xuất Excel PO", use_container_width=True):
                if not selected_rows:
                    st.warning("⚠️ Vui lòng chọn ít nhất một mặt hàng trước khi xuất Excel.")
                else:
                    result_df = pd.DataFrame(selected_rows)
                    buffer = create_excel(result_df)
                    file_name = f"PO_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    st.download_button(
                        label="📥 Tải file Excel",
                        data=buffer,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
